import openpyxl as xl
import xlrd
import os

try:
    filename = "none"
    filename1 = "none"
    # mop sheet search for col location
    p_name = 0
    w_sale = 0
    rtail = 0
    inum = 0
    shipping = 0
    variations = 0
    # clean sheet search for col locations
    s_p_name = 0
    sw_sale = 0
    srtail = 0
    sinum = 0
    s_shipping = 0
    s_duty = 0
    s_variations = 0
    for file in os.listdir('.'):
        if "_CLEAN" in file:
            filename = file
    wb1 = xlrd.open_workbook(filename)
    ws1 = wb1.sheet_by_index(0)
    for file in os.listdir('.'):
        if "template" in file:
            filename1 = file
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active
    mc = ws1.ncols
    mr = ws1.nrows
    tmc = ws2.max_column
    tmr = ws2.max_row
    # here we search for the columns in the clean sheet
    for y in range(0, mc):
        if ws1.cell_value(rowx=14, colx=y) == "Product Name ":
            s_p_name = y
        if ws1.cell_value(rowx=14, colx=y) == "ToMo Wholesale Price (USD)":
            sw_sale = y
        if ws1.cell_value(rowx=14, colx=y) == "Retail Price, MSRP, RRP (USD)":
            srtail = y
        if ws1.cell_value(rowx=14, colx=y) == "Item Number (SKU)":
            sinum = y
        if ws1.cell_value(rowx=14, colx=y) == "Total Estimated Shipping Costs ":
            s_shipping = y
        if ws1.cell_value(rowx=14, colx=y) == "Duties":
            s_duty = y
        if "Variation" in ws1.cell_value(rowx=14, colx=y):
            s_variations = y
    # here we search for the rows in the template
    for p in range(1, tmc):
        for y in range(1, 35):
            if ws2.cell(row=y, column=p).value == "Product Name ":
                p_name = p
            if ws2.cell(row=y, column=p).value == "ToMo Wholesale Price (USD)":
                w_sale = p
            if ws2.cell(row=y, column=p).value == "Retail Price, MSRP, RRP (USD)":
                rtail = p
            if ws2.cell(row=y, column=p).value == "Item Number (SKU)":
                inum = p
            if ws2.cell(row=y, column=p).value == "ToMo Shipping Estimate":
                shipping = p
            if ws2.cell(row=y, column=p).value == "Variation":
                variations = p
    for i in range(18, mr + 1):
        c = ws1.cell_value(rowx=i - 1, colx=sinum)
        ws2.cell(row=i - 16, column=inum).value = c
        c = ws1.cell_value(rowx=i - 1, colx=s_p_name)
        ws2.cell(row=i - 16, column=p_name).value = c
        c = ws1.cell_value(rowx=i - 1, colx=s_variations)
        ws2.cell(row=i - 16, column=variations).value = c
        c = ws1.cell_value(rowx=i - 1, colx=sw_sale)
        ws2.cell(row=i - 16, column=w_sale).value = c
        c = ws1.cell_value(rowx=i - 1, colx=srtail)
        ws2.cell(row=i - 16, column=rtail).value = c
        c = ws1.cell_value(rowx=i - 1, colx=s_shipping)
        if type(ws1.cell_value(rowx=i - 1, colx=s_duty)) is str:
            ws2.cell(row=i - 16, column=shipping).value = c
        else:
            c = c + ws1.cell_value(rowx=i - 1, colx=s_duty)
            ws2.cell(row=i - 16, column=shipping).value = c
    wb2.save(str(filename[:-9] + "MOP.xlsx"))
except Exception as e:
    errorreport = open('error.txt', 'w')
    print(str(e), file=errorreport)
    errorreport.close()
